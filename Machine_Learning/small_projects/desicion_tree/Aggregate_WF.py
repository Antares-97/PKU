import os
import xlrd
import xlwt
from functools import reduce
import xlsxwriter
import openpyxl


stopwords = {'的', '年', '月', '日', '时', '不', '岁', '中', '说', '称', '做', '高', '达', '新', '未', '名'}


def open_excel(file_path):
    file = xlrd.open_workbook(file_path)
    return file


def sign(x):
    if x >= 0:
        return 1
    else:
        return -1


def dictsum(dict):
    num = 0
    for key in dict:
        num += dict[key]
    return num


def excel_process(file_path):
    file = open_excel(file_path)

    Words = set()
    dict0 = {}
    dict1 = {}
    dict2 = {}
    dict3 = {}
    dict4 = {}
    dict5 = {}
    dict6 = {}
    dict7 = {}
    dict8 = {}

    sheet = file.sheet_by_index(0)
    nrows = sheet.nrows
    for rownum in range(0, nrows):
        row = sheet.row_values(rownum)
        word = row[0]
        dict0[word] = row[1]
        if word not in stopwords:
            Words.add(word)
    num0 = dictsum(dict0)
    print(num0)


    sheet = file.sheet_by_index(1)
    nrows = sheet.nrows
    for rownum in range(0, nrows):
        row = sheet.row_values(rownum)
        word = row[0]
        dict1[word] = row[1]
        if word not in stopwords:
            Words.add(word)
    num1 = dictsum(dict1)
    print(num1)


    sheet = file.sheet_by_index(2)
    nrows = sheet.nrows
    for rownum in range(0, nrows):
        row = sheet.row_values(rownum)
        word = row[0]
        dict2[word] = row[1]
        if word not in stopwords:
            Words.add(word)
    num2 = dictsum(dict2)
    print(num2)

    sheet = file.sheet_by_index(3)
    nrows = sheet.nrows
    for rownum in range(0, nrows):
        row = sheet.row_values(rownum)
        word = row[0]
        dict3[word] = row[1]
        if word not in stopwords:
            Words.add(word)
    num3 = dictsum(dict3)
    print(num3)

    sheet = file.sheet_by_index(4)
    nrows = sheet.nrows
    for rownum in range(0, nrows):
        row = sheet.row_values(rownum)
        word = row[0]
        dict4[word] = row[1]
        if word not in stopwords:
            Words.add(word)
    num4 = dictsum(dict4)
    print(num4)

    sheet = file.sheet_by_index(5)
    nrows = sheet.nrows
    for rownum in range(0, nrows):
        row = sheet.row_values(rownum)
        word = row[0]
        dict5[word] = row[1]
        if word not in stopwords:
            Words.add(word)
    num5 = dictsum(dict5)
    print(num5)

    sheet = file.sheet_by_index(6)
    nrows = sheet.nrows
    for rownum in range(0, nrows):
        row = sheet.row_values(rownum)
        word = row[0]
        dict6[word] = row[1]
        if word not in stopwords:
            Words.add(word)
    num6 = dictsum(dict6)
    print(num6)

    sheet = file.sheet_by_index(7)
    nrows = sheet.nrows
    for rownum in range(0, nrows):
        row = sheet.row_values(rownum)
        word = row[0]
        dict7[word] = row[1]
        if word not in stopwords:
            Words.add(word)
    num7 = dictsum(dict7)
    print(num7)

    sheet = file.sheet_by_index(8)
    nrows = sheet.nrows
    for rownum in range(0, nrows):
        row = sheet.row_values(rownum)
        word = row[0]
        dict8[word] = row[1]
        if word not in stopwords:
            Words.add(word)
    num8 = dictsum(dict8)
    print(num8)

    TotalNum = num0 + num1 + num2 + num3 + num4 + num5 + num6 + num7 + num8  # 总频数
    print(TotalNum)


    wordbook = xlwt.Workbook(encoding = 'utf-8')
    wordsheet = wordbook.add_sheet('Aggregate')
    CHI0 = wordbook.add_sheet('Education')
    CHI1 = wordbook.add_sheet('Entertainment')
    CHI2 = wordbook.add_sheet('Finance')
    CHI3 = wordbook.add_sheet('Health')
    CHI4 = wordbook.add_sheet('Military')
    CHI5 = wordbook.add_sheet('RealEstate')
    CHI6 = wordbook.add_sheet('Securities')
    CHI7 = wordbook.add_sheet('Sports')
    CHI8 = wordbook.add_sheet('Tech')
    CHI = wordbook.add_sheet('X^2')

    wordsheet.write(0, 1, 'Education')
    wordsheet.write(0, 2, 'Entertainment')
    wordsheet.write(0, 3, 'Finance')
    wordsheet.write(0, 4, 'Health')
    wordsheet.write(0, 5, 'Military')
    wordsheet.write(0, 6, 'RealEstate')
    wordsheet.write(0, 7, 'Securities')
    wordsheet.write(0, 8, 'Sports')
    wordsheet.write(0, 9, 'Tech')


    n = 1
    for key in Words:
        wordsheet.write(n, 0, key)
        wordsheet.write(n, 1, dict0.get(key, 0))
        wordsheet.write(n, 2, dict1.get(key, 0))
        wordsheet.write(n, 3, dict2.get(key, 0))
        wordsheet.write(n, 4, dict3.get(key, 0))
        wordsheet.write(n, 5, dict4.get(key, 0))
        wordsheet.write(n, 6, dict5.get(key, 0))
        wordsheet.write(n, 7, dict6.get(key, 0))
        wordsheet.write(n, 8, dict7.get(key, 0))
        wordsheet.write(n, 9, dict8.get(key, 0))
        num = dict0.get(key, 0) + dict1.get(key, 0) + dict2.get(key, 0) + dict3.get(key, 0) + dict4.get(key, 0) + dict5.get(key, 0) + dict6.get(key, 0) + dict7.get(key, 0) + dict8.get(key, 0)
        wordsheet.write(n, 10, num)
        n += 1

    n = 0
    for key in Words:
        CHI0.write(n, 0, key)
        CHI.write(n, 0, key)
        num = dict0.get(key, 0) + dict1.get(key, 0) + dict2.get(key, 0) + dict3.get(key, 0) + dict4.get(key, 0) + dict5.get(key, 0) + dict6.get(key, 0) + dict7.get(key, 0) + dict8.get(key, 0)
        n11 = dict0.get(key, 0)  # 可能为零! 因为该词条可能不出现在某一种类当中
        n12 = num - n11
        n21 = num0 - n11
        n22 = TotalNum - num0 - num + n11
        X = sign(n11*n22-n12*n21)*TotalNum*(n11*n22-n12*n21)*(n11*n22-n12*n21)/(n11*n11*n22*n22*n12*n12*n21*n21+1)
        CHI0.write(n, 1, X)
        CHI.write(n, 1, X)
        n += 1

    n = 0
    for key in Words:
        CHI1.write(n, 0, key)
        num = dict0.get(key, 0) + dict1.get(key, 0) + dict2.get(key, 0) + dict3.get(key, 0) + dict4.get(key,
                                                                                                        0) + dict5.get(
            key, 0) + dict6.get(key, 0) + dict7.get(key, 0) + dict8.get(key, 0)
        n11 = dict1.get(key, 0)
        n12 = num - n11
        n21 = num1 - n11
        n22 = TotalNum - num1 - num + n11
        X = sign(n11 * n22 - n12 * n21) * TotalNum * (n11 * n22 - n12 * n21) * (n11 * n22 - n12 * n21) / (
                    n11 * n11 * n22 * n22 * n12 * n12 * n21 * n21+1)  # 分母加一, 来防止分母为0
        CHI1.write(n, 1, X)
        CHI.write(n, 2, X)
        n += 1

    n = 0
    for key in Words:
        CHI2.write(n, 0, key)
        num = dict0.get(key, 0) + dict1.get(key, 0) + dict2.get(key, 0) + dict3.get(key, 0) + dict4.get(key,
                                                                                                        0) + dict5.get(
            key, 0) + dict6.get(key, 0) + dict7.get(key, 0) + dict8.get(key, 0)
        n11 = dict2.get(key, 0)
        n12 = num - n11
        n21 = num2 - n11
        n22 = TotalNum - num2 - num + n11
        X = sign(n11 * n22 - n12 * n21) * TotalNum * (n11 * n22 - n12 * n21) * (n11 * n22 - n12 * n21) / (
                    n11 * n11 * n22 * n22 * n12 * n12 * n21 * n21+1)
        CHI2.write(n, 1, X)
        CHI.write(n, 3, X)
        n += 1

    n = 0
    for key in Words:
        CHI3.write(n, 0, key)
        num = dict0.get(key, 0) + dict1.get(key, 0) + dict2.get(key, 0) + dict3.get(key, 0) + dict4.get(key,
                                                                                                        0) + dict5.get(
            key, 0) + dict6.get(key, 0) + dict7.get(key, 0) + dict8.get(key, 0)
        n11 = dict3.get(key, 0)
        n12 = num - n11
        n21 = num3 - n11
        n22 = TotalNum - num3 - num + n11
        X = sign(n11 * n22 - n12 * n21) * TotalNum * (n11 * n22 - n12 * n21) * (n11 * n22 - n12 * n21) / (
                    n11 * n11 * n22 * n22 * n12 * n12 * n21 * n21+1)
        CHI3.write(n, 1, X)
        CHI.write(n, 4, X)
        n += 1

    n = 0
    for key in Words:
        CHI4.write(n, 0, key)
        num = dict0.get(key, 0) + dict1.get(key, 0) + dict2.get(key, 0) + dict3.get(key, 0) + dict4.get(key,
                                                                                                        0) + dict5.get(
            key, 0) + dict6.get(key, 0) + dict7.get(key, 0) + dict8.get(key, 0)
        n11 = dict4.get(key, 0)
        n12 = num - n11
        n21 = num4 - n11
        n22 = TotalNum - num4 - num + n11
        X = sign(n11 * n22 - n12 * n21) * TotalNum * (n11 * n22 - n12 * n21) * (n11 * n22 - n12 * n21) / (
                    n11 * n11 * n22 * n22 * n12 * n12 * n21 * n21+1)
        CHI4.write(n, 1, X)
        CHI.write(n, 5, X)
        n += 1

    n = 0
    for key in Words:
        CHI5.write(n, 0, key)
        num = dict0.get(key, 0) + dict1.get(key, 0) + dict2.get(key, 0) + dict3.get(key, 0) + dict4.get(key,
                                                                                                        0) + dict5.get(
            key, 0) + dict6.get(key, 0) + dict7.get(key, 0) + dict8.get(key, 0)
        n11 = dict5.get(key, 0)
        n12 = num - n11
        n21 = num5 - n11
        n22 = TotalNum - num5 - num + n11
        X = sign(n11 * n22 - n12 * n21) * TotalNum * (n11 * n22 - n12 * n21) * (n11 * n22 - n12 * n21) / (
                    n11 * n11 * n22 * n22 * n12 * n12 * n21 * n21+1)
        CHI5.write(n, 1, X)
        CHI.write(n, 6, X)
        n += 1

    n = 0
    for key in Words:
        CHI6.write(n, 0, key)
        num = dict0.get(key, 0) + dict1.get(key, 0) + dict2.get(key, 0) + dict3.get(key, 0) + dict4.get(key,
                                                                                                        0) + dict5.get(
            key, 0) + dict6.get(key, 0) + dict7.get(key, 0) + dict8.get(key, 0)
        n11 = dict6.get(key, 0)
        n12 = num - n11
        n21 = num6 - n11
        n22 = TotalNum - num6 - num + n11
        X = sign(n11 * n22 - n12 * n21) * TotalNum * (n11 * n22 - n12 * n21) * (n11 * n22 - n12 * n21) / (
                    n11 * n11 * n22 * n22 * n12 * n12 * n21 * n21+1)
        CHI6.write(n, 1, X)
        CHI.write(n, 7, X)
        n += 1

    n = 0
    for key in Words:
        CHI7.write(n, 0, key)
        num = dict0.get(key, 0) + dict1.get(key, 0) + dict2.get(key, 0) + dict3.get(key, 0) + dict4.get(key,
                                                                                                        0) + dict5.get(
            key, 0) + dict6.get(key, 0) + dict7.get(key, 0) + dict8.get(key, 0)
        n11 = dict7.get(key, 0)
        n12 = num - n11
        n21 = num7 - n11
        n22 = TotalNum - num7 - num + n11
        X = sign(n11 * n22 - n12 * n21) * TotalNum * (n11 * n22 - n12 * n21) * (n11 * n22 - n12 * n21) / (
                    n11 * n11 * n22 * n22 * n12 * n12 * n21 * n21+1)
        CHI7.write(n, 1, X)
        CHI.write(n, 8, X)
        n += 1

    n = 0
    for key in Words:
        CHI8.write(n, 0, key)
        num = dict0.get(key, 0) + dict1.get(key, 0) + dict2.get(key, 0) + dict3.get(key, 0) + dict4.get(key,
                                                                                                        0) + dict5.get(
            key, 0) + dict6.get(key, 0) + dict7.get(key, 0) + dict8.get(key, 0)
        n11 = dict8.get(key, 0)
        n12 = num - n11
        n21 = num8 - n11
        n22 = TotalNum - num8 - num + n11
        X = sign(n11 * n22 - n12 * n21) * TotalNum * (n11 * n22 - n12 * n21) * (n11 * n22 - n12 * n21) / (
                    n11 * n11 * n22 * n22 * n12 * n12 * n21 * n21+1)
        CHI8.write(n, 1, X)
        CHI.write(n, 9, X)
        n += 1


    wordbook.save('WORDS.xls')


def make_sets(file_path):
    wordlist = set()
    file = open_excel(file_path)

    sheet = file.sheet_by_index(1)
    for rownum in range(0, 200):
        row = sheet.row_values(rownum)
        word = row[0]
        wordlist.add(word)

    sheet = file.sheet_by_index(2)
    for rownum in range(0, 200):
        row = sheet.row_values(rownum)
        word = row[0]
        wordlist.add(word)

    sheet = file.sheet_by_index(3)
    for rownum in range(0, 200):
        row = sheet.row_values(rownum)
        word = row[0]
        wordlist.add(word)

    sheet = file.sheet_by_index(4)
    for rownum in range(0, 200):
        row = sheet.row_values(rownum)
        word = row[0]
        wordlist.add(word)

    sheet = file.sheet_by_index(5)
    for rownum in range(0, 200):
        row = sheet.row_values(rownum)
        word = row[0]
        wordlist.add(word)

    sheet = file.sheet_by_index(6)
    for rownum in range(0, 200):
        row = sheet.row_values(rownum)
        word = row[0]
        wordlist.add(word)

    sheet = file.sheet_by_index(7)
    for rownum in range(0, 200):
        row = sheet.row_values(rownum)
        word = row[0]
        wordlist.add(word)

    sheet = file.sheet_by_index(8)
    for rownum in range(0, 200):
        row = sheet.row_values(rownum)
        word = row[0]
        wordlist.add(word)

    sheet = file.sheet_by_index(9)
    for rownum in range(0, 200):
        row = sheet.row_values(rownum)
        word = row[0]
        wordlist.add(word)

    wordbook = xlsxwriter.Workbook('wordlist.xlsx')
    wordsheet = wordbook.add_worksheet('wordlist')
    j = 0
    for key in wordlist:
        print(key)
        wordsheet.write(j, 0, key)
        j += 1


def make_Train_Sets(wdl_path, path):
    wordlist = set()
    file = xlrd.open_workbook(wdl_path)
    File = open('/Users/macbookair/iCloud/Desktop/Daily/Second Major/CS/Machine Learning/hw1/trainsets.txt', 'w', encoding='utf-8')

    sheet = file.sheet_by_index(0)
    for rownum in range(0, 1799):
        row = sheet.row_values(rownum)
        word = row[0]
        wordlist.add(word)

    trainset = xlsxwriter.Workbook('TrainSets.xlsx')
    TrainSets = trainset.add_worksheet('TrainSets')
    j = 0
    for key in wordlist:
        TrainSets.write(0, j, key)
        j += 1
        s = '{0}\t'.format(key)
        File.write(s)
    TrainSets.write(0, j, 'lables:')
    s = '{0}\n'.format('lables:')
    File.write(s)

    i = 1
    j = 0
    for catepath, category, filename in os.walk(path, False):
        try:
            filename.remove('.DS_Store')
        except:
            pass
        filelist = filename
        for file in filelist:
            FILE = open(catepath + '/' + file, 'r').read()
            WORDS = set()
            word = ''
            for char in FILE:
                if char != '\t':  # 工作原理: 一个字符一个字符地读, 将不是制表符的字符丢进 word 字符串当中
                    word += char  # 连接字符
                else:  # 读到了制表符, 就从 word 字符串当中读取一个单词, 进行词频统计!
                    WORDS.add(word)
                    word = ''  # 清空字符串,等待下一个词的输入

            for key in wordlist:
                if key in WORDS:
                    TrainSets.write(i, j, 'YES')
                    s = 'YES\t'
                    File.write(s)
                else:
                    TrainSets.write(i, j, 'NO')
                    s = 'NO\t'
                    File.write(s)
                j += 1
            TrainSets.write(i, j, str(category))  # 最后一个 label 为种类
            s = '{0}\n'.format(str(category))
            File.write(s)
            WORDS.clear()
            i += 1
            j = 0  # 换行, 回车


if __name__ == "__main__":
    file_path = '/Users/macbookair/iCloud/Desktop/Daily/Second Major/CS/Machine Learning/hw1/WORDS.xlsx'
    wdl_path = '/Users/macbookair/iCloud/Desktop/Daily/Second Major/CS/Machine Learning/hw1/wordlist.xlsx'
    path = '/Users/macbookair/iCloud/Desktop/Daily/Second Major/CS/Machine Learning/hw1/new_weibo_13638/Train_Set/'
    make_sets(file_path)
    # make_Train_Sets(wdl_path, path)


